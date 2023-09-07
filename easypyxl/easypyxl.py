"""
26 Sep 2021
Myung-Joon Kwon (CauchyComplete)
corundum240@gmail.com
"""
import openpyxl
import os
import collections.abc
import time


class Workbook:
    def __init__(self, excel_filepath, verbose=True, backup=True):
        if '.' not in excel_filepath:
            excel_filepath = excel_filepath + ".xlsx"
        else:
            if os.path.splitext(excel_filepath)[-1] != ".xlsx":
                raise IOError(f"EasyPyXL only supports '.xlsx' file. Given: {excel_filepath}")
        need_to_save = False
        if os.path.isfile(excel_filepath):
            workbook = openpyxl.load_workbook(excel_filepath, read_only=False)
            if verbose:
                print(f"EasyPyXL loaded workbook: {excel_filepath}")
            self.empty_file = False
            if backup:
                filename = os.path.splitext(excel_filepath)[0] + "_easypyxl_backup"
                for i in range(0, 10000000):
                    new_filepath = f"{filename}_{i}.xlsx"
                    if not os.path.isfile(new_filepath):
                        break
                else:
                    raise IOError(f"EasyPyXL cannot backup: {excel_filepath}")
                if verbose:
                    print(f"EasyPyXL backup file: {new_filepath}")
                workbook.save(new_filepath)
        else:
            workbook = openpyxl.Workbook()
            if verbose:
                print(f"EasyPyXL created workbook: {excel_filepath}")
            self.empty_file = True
            backup = False
            need_to_save = True

        self.workbook = workbook
        self.excel_filepath = excel_filepath
        self.verbose = verbose
        self.saved_error_counter = 0
        self.backup = backup
        self.prev_saved_time = time.time()
        if need_to_save is True:
            self.save_excel()

    class Cursor:
        def __init__(self, workbook_class, sheet, start_cell, seq_len, move_vertical, reader, auto_save,
                     auto_save_time):
            self.workbook_class = workbook_class
            self.sheet = sheet
            self.start_cell = start_cell
            self.move_vertical = move_vertical
            self.item_count = 0
            self.seq_len = seq_len
            self.reader = reader
            self.auto_save = auto_save
            self.auto_save_time = auto_save_time

        def _write_cell(self, val):
            if self.move_vertical:
                self.sheet.cell(self.start_cell[0] + (self.item_count % self.seq_len),
                                self.start_cell[1] + (self.item_count // self.seq_len)).value = val
            else:
                self.sheet.cell(self.start_cell[0] + (self.item_count // self.seq_len),
                                self.start_cell[1] + (self.item_count % self.seq_len)).value = val
            self.item_count += 1

        def write_cell(self, val):
            if self.reader:
                raise PermissionError("EasyPyXL: You cannot write_cell() using a cursor with reader=True")
            if isinstance(val, collections.abc.Sequence) and not isinstance(val, str):
                # list or tuple
                for v in val:
                    self._write_cell(v)
            else:
                # string, int, or float
                self._write_cell(val)

            if self.auto_save:
                self.workbook_class.save_excel(self.auto_save_time)

        def _read_cell(self):
            if self.move_vertical:
                val = self.sheet.cell(self.start_cell[0] + (self.item_count % self.seq_len),
                                      self.start_cell[1] + (self.item_count // self.seq_len)).value
            else:
                val = self.sheet.cell(self.start_cell[0] + (self.item_count // self.seq_len),
                                      self.start_cell[1] + (self.item_count % self.seq_len)).value
            self.item_count += 1
            return val

        def read_cell(self, amount=1):
            if not self.reader:
                raise PermissionError("EasyPyXL: You cannot read_cell() using a cursor with reader=False")
            if amount >= 2:
                result = [self._read_cell() for _ in range(amount)]
            elif amount <= 0:
                raise ValueError("EasyPyXL: Cannot read_cell() - amount must be a positive integer.")
            else:
                result = self._read_cell()
            return result

        def read_line(self, amount=1):
            if not self.reader:
                raise PermissionError("EasyPyXL: You cannot read_line() using a cursor with reader=False")
            if amount >= 2:
                result = [self.read_cell(self.seq_len) for _ in range(amount)]
            elif amount <= 0:
                raise ValueError("EasyPyXL: Cannot read_cell() - amount must be a positive integer.")
            else:
                result = self.read_cell(self.seq_len)
            return result

        def skip_cell(self, amount=1):
            self.item_count += amount

        def skip_line(self, amount=1):
            self.item_count += amount * self.seq_len

    def new_cursor(self, sheetname, start_cell, seq_len, move_vertical=False, overwrite=False, reader=False,
                   auto_save=True, auto_save_time=0):
        if auto_save is False and auto_save_time != 0:
            raise ValueError("auto_save is False but auto_save_time is given.")
        if isinstance(start_cell, str):
            start_cell = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
        if self.empty_file:
            sheet = self.workbook.active
            sheet.title = sheetname
            self.empty_file = False
            if self.verbose:
                print(f"EasyPyXL created sheet: {sheetname}")
        else:
            if sheetname is None:
                sheet = self.workbook.active
                sheetname = sheet.title
                if self.verbose:
                    print(f"EasyPyXL loaded active sheet: {sheetname}")
            elif sheetname in self.workbook.sheetnames:
                sheet = self.workbook[sheetname]
                if self.verbose:
                    print(f"EasyPyXL loaded sheet: {sheetname}")
            else:
                sheet = self.workbook.create_sheet(sheetname)
                if self.verbose:
                    print(f"EasyPyXL created sheet: {sheetname}")
        prev_cell_value = sheet.cell(*start_cell).value
        if prev_cell_value is not None and not overwrite and not reader:
            raise ValueError(f"EasyPyXL: start_cell {start_cell} of '{sheetname}' is not empty! "
                             f"Current value: {str(sheet.cell(*start_cell).value)}. To overwrite, set overwrite=True.")
        cursor = self.Cursor(self, sheet, start_cell, seq_len, move_vertical, reader, auto_save, auto_save_time)
        return cursor

    def save_excel(self, auto_save_time=0):
        if auto_save_time == 0:
            self._save_excel()
        elif time.time() - self.prev_saved_time > auto_save_time:
            self._save_excel()

    def _save_excel(self):
        try:
            self.workbook.save(self.excel_filepath)
        except PermissionError:
            filename = os.path.splitext(self.excel_filepath)[0]
            if self.saved_error_counter > 0:
                filename = '_'.join(filename.split('_')[:-1])
            for i in range(self.saved_error_counter, 10000000):
                new_filepath = f"{filename}_{i}.xlsx"
                if not os.path.isfile(new_filepath):
                    self.saved_error_counter += 1
                    break
            else:
                raise IOError(f"EasyPyXL cannot save file: {self.excel_filepath}")
            self.excel_filepath = new_filepath
            if self.verbose:
                print(f"EasyPyXL saved to new file: {self.excel_filepath}")
            self.workbook.save(self.excel_filepath)
        self.prev_saved_time = time.time()

    class SmartCursor:
        def __init__(self, workbook_class, sheet, start_cell, row_names, col_names, corner_name, auto_save, auto_save_time):
            self.workbook_class = workbook_class
            self.sheet = sheet
            self.start_cell = start_cell
            self.auto_save = auto_save
            self.auto_save_time = auto_save_time

            # Check corner_name
            if corner_name is not None:
                if sheet.cell(*start_cell).value is not None:
                    if not str(sheet.cell(*start_cell).value) == str(corner_name):
                        raise ValueError(f"corner_name is given ({corner_name}) but not equal to the value in Excel ({sheet.cell(*start_cell).value}).")
                sheet.cell(*start_cell).value = str(corner_name)


            # initial row_names check
            excel_row_names = []
            i = 1
            while True:
                cur_cell_val = self.sheet.cell(self.start_cell[0] + i, self.start_cell[1]).value
                if cur_cell_val is not None:
                    excel_row_names.append(str(cur_cell_val))
                    i += 1
                else:
                    break
            if row_names is None:
                self.row_names = excel_row_names
            elif len(excel_row_names) == 0:
                self.row_names = row_names
                for i, row_name in enumerate(self.row_names):
                    self.sheet.cell(self.start_cell[0] + 1 + i, self.start_cell[1]).value = row_name
            else:
                if len(row_names) == len(excel_row_names) and all(
                        [str(x) == str(y) for x, y in zip(row_names, excel_row_names)]):
                    self.row_names = row_names
                    for i, row_name in enumerate(self.row_names):
                        self.sheet.cell(self.start_cell[0] + 1 + i, self.start_cell[1]).value = row_name
                else:
                    raise ValueError("`row_names` is given but does not match the values in Excel.")
            self.row_names = [str(x) for x in self.row_names]

            # initial col_names check
            excel_col_names = []
            i = 1
            while True:
                cur_cell_val = self.sheet.cell(self.start_cell[0], self.start_cell[1] + i).value
                if cur_cell_val is not None:
                    excel_col_names.append(str(cur_cell_val))
                    i += 1
                else:
                    break
            if col_names is None:
                self.col_names = excel_col_names
            elif len(excel_col_names) == 0:
                self.col_names = col_names
                for i, col_names in enumerate(self.col_names):
                    self.sheet.cell(self.start_cell[0], self.start_cell[1] + 1 + i).value = col_names
            else:
                if len(col_names) == len(excel_col_names) and all(
                        [str(x) == str(y) for x, y in zip(col_names, excel_col_names)]):
                    self.col_names = col_names
                    for i, col_names in enumerate(self.col_names):
                        self.sheet.cell(self.start_cell[0], self.start_cell[1] + 1 + i).value = col_names
                else:
                    raise ValueError("`col_names` is given but does not match the values in Excel.")
            self.col_names = [str(x) for x in self.col_names]

        def _write_cell(self, row_num, col_num, val):
            self.sheet.cell(row_num, col_num).value = val

        def write_cell(self, row_name, col_name, val):
            if row_name not in self.row_names:
                self.row_names.append(str(row_name))
                self._write_cell(self.start_cell[0] + len(self.row_names), self.start_cell[1], row_name)
            if col_name not in self.col_names:
                self.col_names.append(str(col_name))
                self._write_cell(self.start_cell[0], self.start_cell[1] + len(self.col_names), col_name)
            row_num = self.start_cell[0] + 1 + self.row_names.index(str(row_name))
            col_num = self.start_cell[1] + 1 + self.col_names.index(str(col_name))
            self._write_cell(row_num, col_num, val)

            if self.auto_save:
                self.workbook_class.save_excel(self.auto_save_time)

        def _read_cell(self, row_num, col_num):
            return self.sheet.cell(row_num, col_num).value

        def read_cell(self, row_name, col_name):
            if row_name not in self.row_names or col_name not in self.col_names:
                return None
            row_num = self.start_cell[0] + 1 + self.row_names.index(str(row_name))
            col_num = self.start_cell[1] + 1 + self.col_names.index(str(col_name))
            return self._read_cell(row_num, col_num)

    def new_smart_cursor(self, sheetname, start_cell, row_names=None, col_names=None, corner_name=None, auto_save=True, auto_save_time=0):
        if auto_save is False and auto_save_time != 0:
            raise ValueError("auto_save is False but auto_save_time is given.")
        if row_names:
            assert isinstance(row_names, list)
            if not len(row_names) == len(set(row_names)):
                raise ValueError("row_names must not have duplicate elements")
        if col_names:
            assert isinstance(col_names, list)
            if not len(col_names) == len(set(col_names)):
                raise ValueError("row_names must not have duplicate elements")
        if isinstance(start_cell, str):
            start_cell = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
        if self.empty_file:
            sheet = self.workbook.active
            sheet.title = sheetname
            self.empty_file = False
            if self.verbose:
                print(f"EasyPyXL created sheet: {sheetname}")
        else:
            if sheetname is None:
                sheet = self.workbook.active
                sheetname = sheet.title
                if self.verbose:
                    print(f"EasyPyXL loaded active sheet: {sheetname}")
            elif sheetname in self.workbook.sheetnames:
                sheet = self.workbook[sheetname]
                if self.verbose:
                    print(f"EasyPyXL loaded sheet: {sheetname}")
            else:
                sheet = self.workbook.create_sheet(sheetname)
                if self.verbose:
                    print(f"EasyPyXL created sheet: {sheetname}")
        smart_cursor = self.SmartCursor(self, sheet, start_cell, row_names, col_names, corner_name, auto_save, auto_save_time)
        return smart_cursor
